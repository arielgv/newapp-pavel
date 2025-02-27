import re
from typing import Callable, Set, Dict, Any, Optional, Tuple
from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from app.utils.logger import logger
import io

# Constants for error messages
VENDOR_NOT_FOUND = "Vendor not found"
ENTITY_NOT_FOUND = "Entity not found at Database"
CONTAINER_NOT_FOUND = "Container number not found"
COMMENT_AUTHOR = "COSA Validation System"
COMMENT_SEPARATOR = ", "

def simple_slugify(text):
    """Convert text to a simple slug format (lowercase, hyphenated)"""
    if text is None:
        return ''
    text = str(text).lower()
    text = re.sub(r'[^\w\s-]', '', text)
    return re.sub(r'[-\s]+', '-', text).strip('-_')

def normalize_container_number(number):
    """Normalize container numbers by removing spaces and converting to uppercase"""
    if number is None:
        return ''
    # Normalize by removing spaces and converting to uppercase
    return re.sub(r'\s', '', str(number).upper())

class ValidationStats:
    """Tracks validation statistics and errors"""
    def __init__(self):
        self.total_rows = 0
        self.valid_rows = 0
        self.invalid_rows = 0
        self.vendor_not_found = 0
        self.container_not_found = 0
        self.entity_not_found = 0
        self.multiple_errors = set()  # Set of row indices with multiple errors
        self.error_rows = set()  # Set of row indices with any errors

class ExcelValidator:
    """Excel validator for Mother Parkers specific workbooks"""
    
    def __init__(self):
        self.stats = ValidationStats()
    
    def validate_workbook_bytes(self, file_content: bytes) -> Tuple[bytes, Dict[str, Any], Workbook]:
        """
        Validate Excel workbook from bytes content
        
        Args:
            file_content: Bytes content of the Excel file
            
        Returns:
            Tuple containing:
            - The processed workbook as bytes
            - Validation report dictionary
            - Original workbook object for database processing
        """
        # Load workbook from bytes
        wb = load_workbook(io.BytesIO(file_content))
        logger.info(f"Excel file loaded. Available sheets: {wb.sheetnames}")
        
        # Reset statistics
        self.stats = ValidationStats()
        
        # Check if this is the expected Mother Parkers format
        required_sheets = ["Manual Sheet", "Single Supplier Table", "Worksheet- Coffee", "Worksheet- Tea"]
        for sheet in required_sheets:
            if sheet not in wb.sheetnames:
                logger.warning(f"Required sheet '{sheet}' not found. This may not be a Mother Parkers Excel file.")
                return file_content, {"error": f"Required sheet '{sheet}' not found"}, wb
        
        # Get total rows for statistics
        manual_sheet = wb["Manual Sheet"]
        self.stats.total_rows = manual_sheet.max_row - 1  # Exclude header
        
        # Run validations
        wb = self.validate_vendor(wb)
        wb = self.validate_container_number(wb)
        wb = self.validate_entities(wb)
        wb = self.create_manual_sheet_entries(wb)
        
        # Calculate valid rows
        self.stats.valid_rows = self.stats.total_rows - len(self.stats.error_rows)
        self.stats.invalid_rows = len(self.stats.error_rows)
        
        # Generate report
        report = self.generate_validation_report()
        
        # Save workbook to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue(), report, wb

    def validate_vendor(self, wb: Workbook) -> Workbook:
        """Validate vendor names against Single Supplier Table"""
        logger.info("Starting vendor validation...")
        single_supplier_ws = wb["Single Supplier Table"]
        column_name = "Company Name"

        logger.info(f"Validating vendors in 'Single Supplier Table' based on '{column_name}'...")

        company_names = self.get_column_values(single_supplier_ws, column_name, simple_slugify)
        logger.info(f"Company names found: {len(company_names)}")

        if not company_names:
            logger.warning("No company names found.")
            return wb

        manual_sheet_ws = wb["Manual Sheet"]
        column_name = "Exporter Name"
        exporter_name_cell = self.get_column_cell(manual_sheet_ws, column_name)

        if exporter_name_cell is None:
            logger.warning(f"Column '{column_name}' not found in '{manual_sheet_ws.title}'")
            return wb

        logger.info("Validating exporter names...")
        for row in manual_sheet_ws.iter_rows(
            min_col=exporter_name_cell.column,
            max_col=exporter_name_cell.column,
            min_row=exporter_name_cell.row + 1,
            max_row=manual_sheet_ws.max_row,
        ):
            if row[0]:
                cell = row[0]
                if simple_slugify(cell.value) not in company_names:
                    logger.info(f"Vendor not found: {cell.value}")
                    cell = self.mark_cell(cell, cell_comment=VENDOR_NOT_FOUND)
                    self.stats.vendor_not_found += 1
                    self.stats.error_rows.add(cell.row)
                else:
                    cell = self.remove_cell_comment(cell, comment=VENDOR_NOT_FOUND)

        logger.info("Vendor validation completed.")
        return wb

    def validate_container_number(self, wb: Workbook) -> Workbook:
        """Validate container numbers against Coffee and Tea worksheets"""
        manual_sheet_ws = wb["Manual Sheet"]
        manual_sheet_column_name = "Container Number"
        logger.info(f"Validating container numbers in 'Manual Sheet' based on '{manual_sheet_column_name}'...")
        container_number_cell = self.get_column_cell(manual_sheet_ws, manual_sheet_column_name)

        if container_number_cell is None:
            logger.warning(f"Column '{manual_sheet_column_name}' not found in 'Manual Sheet'")
            return wb

        coffee_ws = wb["Worksheet- Coffee"]
        tea_ws = wb["Worksheet- Tea"]
        container_column_name = "Container #"

        container_numbers = set()
        for ws in [coffee_ws, tea_ws]:
            container_numbers.update(self.get_column_values(ws, container_column_name, normalize_container_number))

        if not container_numbers:
            logger.warning("No container numbers found in coffee and tea worksheets.")
            return wb

        logger.info(f"Container numbers found: {len(container_numbers)}")

        not_found_containers = set()
        for row in manual_sheet_ws.iter_rows(
            min_col=container_number_cell.column,
            max_col=container_number_cell.column,
            min_row=container_number_cell.row + 1,
            max_row=manual_sheet_ws.max_row,
        ):
            if row[0]:
                cell = row[0]
                normalized_value = normalize_container_number(str(cell.value))
                if normalized_value not in container_numbers:
                    not_found_containers.add(cell.value)
                    cell = self.mark_cell(cell, cell_comment=CONTAINER_NOT_FOUND)
                    self.stats.container_not_found += 1
                    self.stats.error_rows.add(cell.row)
                else:
                    cell = self.remove_cell_comment(cell, comment=CONTAINER_NOT_FOUND)

        logger.info(f"Container numbers not found: {len(not_found_containers)}")
        logger.info("Container number validation completed.")
        return wb

    def validate_entities(self, wb: Workbook) -> Workbook:
        """Validate entities against Database sheets"""
        logger.info("Starting entity validation...")

        database_others_ws = wb["Database - Others"]
        column_name = "Company Name"

        company_names_db_other = self.get_column_values(
            database_others_ws, column_name, simple_slugify
        )

        database_coop_ws = wb["Database-RA+FT Coop"]
        company_names_db_coop = self.get_column_values(
            database_coop_ws, column_name, simple_slugify
        )

        if not company_names_db_other and not company_names_db_coop:
            logger.warning("No company names found in database sheets.")
            return wb

        logger.info(f"Company names found - Others: {len(company_names_db_other)}, Coop: {len(company_names_db_coop)}")

        manual_sheet_ws = wb["Manual Sheet"]
        target_columns = ["Exporter Name", "Mill Name"]
        for target_column in target_columns:
            logger.info(
                f"Verifying '{target_column}' values against 'Company Name' in database sheets..."
            )
            exporter_name_cell = self.get_column_cell(manual_sheet_ws, target_column)

            if exporter_name_cell is None:
                logger.warning(f"Column '{target_column}' not found in 'Manual Sheet'")
                continue

            for row in manual_sheet_ws.iter_rows(
                min_col=exporter_name_cell.column,
                max_col=exporter_name_cell.column,
                min_row=exporter_name_cell.row + 1,
                max_row=manual_sheet_ws.max_row,
            ):
                if row[0]:
                    cell = row[0]
                    slugified_value = simple_slugify(cell.value)
                    if (
                        slugified_value not in company_names_db_coop
                        and slugified_value not in company_names_db_other
                    ):
                        logger.info(f"Entity not found: {cell.value}")
                        cell = self.mark_cell(cell, cell_comment=ENTITY_NOT_FOUND)
                        self.stats.entity_not_found += 1
                        self.stats.error_rows.add(cell.row)
                        if cell.row in self.stats.error_rows:
                            self.stats.multiple_errors.add(cell.row)
                    else:
                        cell = self.remove_cell_comment(cell, comment=ENTITY_NOT_FOUND)

        logger.info("Entity validation completed.")
        return wb

    def create_manual_sheet_entries(self, wb: Workbook) -> Workbook:
        """Expands rows in the Manual Sheet based on comma-separated values in Coop ID column"""
        sheet_name = "Manual Sheet"
        column_name = "Coop ID"
        logger.info(f"Creating new rows in '{sheet_name}' based on '{column_name}'...")
        ws = wb[sheet_name]
        new_rows = []

        # Find the column index
        column_index = None
        for col_index, col_cells in enumerate(ws.iter_cols(1, ws.max_column), start=1):
            for col_cell in col_cells:
                if col_cell.value == column_name:
                    column_index = col_index
                    break

        # Process each row
        for row in ws.iter_rows(values_only=True):
            # Check if the "Coop ID" column contains a comma
            if column_index and row[column_index - 1] and "," in str(row[column_index - 1]):
                # Split the values in the "Coop ID" column
                values = row[column_index - 1].split(",")
                # Keep the first value as the original row's "Coop ID"
                new_rows.append(list(row))
                new_rows[-1][column_index - 1] = values[0].strip()
                # Create new rows for the remaining values
                for value in values[1:]:
                    # Copy all columns except the specified column
                    new_row = list(row)
                    # Update the "Coop ID" in the new row
                    new_row[column_index - 1] = value.strip()
                    new_rows.append(new_row)
            else:
                # If no comma is found, append the original row as is
                new_rows.append(list(row))

        # Add the expanded rows to the sheet
        for i, row in enumerate(new_rows):
            for j, cell_value in enumerate(row):
                ws.cell(row=i + 1, column=j + 1).value = cell_value
                ws.cell(row=i + 1, column=j + 1).alignment = Alignment(horizontal="left")

        return wb

    def generate_validation_report(self) -> Dict[str, Any]:
        """Generate a validation report dictionary"""
        valid_percentage = (self.stats.valid_rows / self.stats.total_rows) * 100 if self.stats.total_rows > 0 else 0
        total_transactions = self.stats.valid_rows * 2  # 2 transactions per valid record
        
        report = {
            "stats": {
                "total_rows": self.stats.total_rows,
                "valid_rows": self.stats.valid_rows,
                "invalid_rows": self.stats.invalid_rows,
                "valid_percentage": round(valid_percentage, 2)
            },
            "errors": {
                "container_numbers_not_found": self.stats.container_not_found,
                "vendors_not_found": self.stats.vendor_not_found,
                "entities_not_found": self.stats.entity_not_found,
                "rows_with_multiple_errors": len(self.stats.multiple_errors)
            },
            "projections": {
                "valid_records_to_process": self.stats.valid_rows,
                "total_transactions_to_create": total_transactions
            }
        }
        
        return report

    def mark_cell(self, cell: Cell, cell_comment: str, color: str = "FF0000") -> Cell:
        """Mark a cell with provided color and comment"""
        # Fill the cell with provided color
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        # First check if there is any comment already added
        if cell.comment is not None:
            comments = cell.comment.text.split(COMMENT_SEPARATOR)
            comments.append(cell_comment)
            cell.comment = Comment(
                text=COMMENT_SEPARATOR.join(set(comments)), author=COMMENT_AUTHOR
            )
        else:
            cell.comment = Comment(text=cell_comment, author=COMMENT_AUTHOR)

        return cell

    def remove_cell_comment(self, cell: Cell, comment: str) -> Cell:
        """Remove a specific comment from a cell"""
        # Check if current cell has any comment
        if cell.comment is None:
            cell.fill = PatternFill()
            return cell
        else:
            comments = [
                com for com in cell.comment.text.split(COMMENT_SEPARATOR) if com != comment
            ]
            # Check if some comment left
            if comments:
                cell.comment = Comment(
                    text=COMMENT_SEPARATOR.join(set(comments)), author=COMMENT_AUTHOR
                )
            else:
                cell.comment = None
                cell.fill = PatternFill()

            return cell

    def get_column_cell(self, ws: Worksheet, column_name: str) -> Optional[Cell]:
        """Find the cell containing the column header"""
        for row in ws.iter_rows(min_row=1):
            for cell in row:
                if cell.value == column_name:
                    return cell
        return None

    def get_column_values(self, ws: Worksheet, column_name: str, 
                         function: Optional[Callable[[str], str]] = None) -> Set[str]:
        """Get all values from a specific column"""
        cell = self.get_column_cell(ws, column_name)

        if cell is None:
            logger.warning(f"Column '{column_name}' not found in sheet '{ws.title}'")
            return set()

        values = set()
        for row in ws.iter_rows(
            min_col=cell.column,
            max_col=cell.column,
            min_row=cell.row + 1,
            max_row=ws.max_row,
            values_only=True,
        ):
            if row[0]:
                value = function(row[0]) if function else row[0]
                values.add(value)

        return values