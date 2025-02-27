import os
from datetime import datetime
from google.cloud import storage
from app.utils.logger import logger

def backup_file_to_gcs(bucket_name: str, file_path: str, file_content: bytes) -> str:
    """
    Creates a backup of the file in a 'backups' directory in the same GCS bucket.
    
    Args:
        bucket_name: Name of the GCS bucket
        file_path: Path of the file in the bucket
        file_content: File content as bytes
    
    Returns:
        Backup file path
    """
    try:
        # Create backup path
        file_name = os.path.basename(file_path)
        file_base, file_ext = os.path.splitext(file_name)
        timestamp = datetime.now().strftime('%Y_%m_%d_%H_%M_%S')
        backup_name = f"{file_base}_{timestamp}{file_ext}"
        backup_path = f"backups/{backup_name}"
        
        # Upload backup to GCS
        client = storage.Client()
        bucket = client.bucket(bucket_name)
        blob = bucket.blob(backup_path)
        blob.upload_from_string(file_content)
        
        logger.info(f"Created backup at gs://{bucket_name}/{backup_path}")
        return backup_path
    except Exception as e:
        logger.error(f"Failed to create backup: {e}")
        return ""

def is_excel_file(file_path: str) -> bool:
    """
    Check if a file is an Excel file based on its extension.
    
    Args:
        file_path: Path of the file
    
    Returns:
        True if the file is an Excel file, False otherwise
    """
    return file_path.lower().endswith((".xlsx", ".xls"))

def is_mother_parkers_format(file_content: bytes) -> bool:
    """
    Check if the Excel file follows the Mother Parkers format by having the required sheets.
    
    Args:
        file_content: File content as bytes
    
    Returns:
        True if the file has the required sheets, False otherwise
    """
    try:
        import io
        from openpyxl import load_workbook
        
        wb = load_workbook(io.BytesIO(file_content))
        required_sheets = [
            "Manual Sheet", 
            "Single Supplier Table", 
            "Worksheet- Coffee", 
            "Worksheet- Tea"
        ]
        
        # Check if all required sheets exist
        for sheet in required_sheets:
            if sheet not in wb.sheetnames:
                logger.info(f"Missing sheet '{sheet}', not a Mother Parkers format Excel file")
                return False
        
        logger.info("File identified as Mother Parkers format Excel file")
        return True
    except Exception as e:
        logger.error(f"Error checking Excel format: {e}")
        return False
